from django.contrib import admin
from django.contrib.auth.hashers import make_password
from django.conf import settings
from django.contrib.auth.models import User
from django.contrib import messages
import os
import openpyxl  # 代入openpyxl库
from score.models import Score
from student.models import Student
from uploadfile.models import FileUpload

class FileUploadAdmin(admin.ModelAdmin):
    #  配置展示列表，在User版块下的列表展示
    list_display = ('file_name',)
    # 设置只读字段，不允许更改
    readonly_fields = ('teacher',)

    def save_model(self, request, obj, form, change):
        obj.teacher_id = request.user.id # 获取当前老师的ID
        # 调用父类方法保存
        super().save_model(request, obj, form, change)
        # 拼接目录
        file_path = os.path.join(settings.MEDIA_ROOT,obj.file_name.name)
        if request.POST['file_type'] == '1':  # 上传学生信息
            repetition = self.upload_student(file_path,request.user.id)
        elif request.POST['file_type'] == '2':  # 上传成绩信息
            repetition = self.upload_score(file_path)
        # 提示重复数据条数
        if repetition:
            messages.add_message(request, messages.INFO, f'过滤{repetition}条重复数据')
        return

    def upload_student(self,file_path,teacher_id):
        wb = openpyxl.load_workbook(file_path)  # 打开excel
        ws = wb.active  # 选中第一个sheet
        rows = ws.max_row  # 获取行数
        columns = ws.max_column  # 获取列数
        student_list = []
        repetition = 0
        # 从第2行开始遍历每行
        for row in ws.iter_rows(min_row=2, min_col=1, max_row=rows, max_col=columns):
            data = [i.value for i in row]  # 获取每一行数据
            # 去除重复数据
            if Student.objects.filter(student_num=data[0]).exists():
                repetition += 1
                continue
            # 写入User表
            user = User(
                username = data[0], # 以学号作为用户名，防止重复
                password = make_password(settings.STUDENT_INIT_PASSWORD),
            )
            user.save() # 存入数据库
            # 写入student表
            student = Student(
                student_num = data[0],
                name = data[1].strip(), # 去除空格
                gender = 'male' if data[2] == "男" else "femal",
                phone = data[4],
                birthday = data[3],
                user_id = user.id,
                teacher_id = teacher_id
            )
            student_list.append(student)
        Student.objects.bulk_create(student_list) # 批量加入Student表
        return repetition

    def upload_score(self,file_path):
        wb = openpyxl.load_workbook(file_path)  # 打开excel
        ws = wb.active  # 选中第一个sheet
        rows = ws.max_row  # 获取行数
        columns = ws.max_column  # 获取列数
        score_list = []
        repetition = 0
        # 从第2行开始遍历每行
        for row in ws.iter_rows(min_row=2, min_col=1, max_row=rows, max_col=columns):
            data = [i.value for i in row]  # 获取每一行数据
            # 查找student表，获取student_id
            try:
                student = Student.objects.get(student_num=data[1])
            except:
                continue
            # 去除重复数据
            if Score.objects.filter(title=data[0],student_id=student.id).exists():
                repetition += 1
                continue
            # 写入student表
            score = Score(
                title = data[0],         # 标题
                student_id = student.id, # 学生id
                score = data[-1]         # 学生分数
            )
            score_list.append(score)

        Score.objects.bulk_create(score_list) # 批量加入Score表
        return repetition

admin.site.register(FileUpload, FileUploadAdmin)